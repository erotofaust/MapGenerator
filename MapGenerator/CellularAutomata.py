import numpy as np
import random
from openpyxl import Workbook
import matplotlib.pyplot as plt

# openpyxl을 이용해 엑셀로 데이터 저장
# matplotlib 이용해 맵의 형태를 히트맵으로 표현

map_x_max = 200
map_y_max = 100
map_size = [map_x_max,int(map_y_max/2)]

# map의 사이즈 지정
# 이때 좌표계 특성에 따라 총 타일 수는 map_x_max*map_y_max*0.5
# map_x_max, map_y_max는 반드시 짝수여야 함

cell_data = ["x","y","WorldMapTile_Id", "WorldMapTerrain_Id"]
grass_prob = 0.41
cellular_iter_num = 5
lamb = 0.02


shadowTile_id = 20
waterTile_id = [1,1,1,
                1,1,1,
                1,1,1]
snowTile_id = 16
grassTile_id = 6

woodTerrain_id = 5
stoneTerrain_id = 15

# 최초 세팅값
# water_id는 9-slice
# 0 1 2
# 3 4 5
# 6 7 8

map = np.zeros((map_size[0], map_size[1], len(cell_data)))


for n in range(0,map_size[0]):
    for m in range(0, map_size[1]):
        a = random.random()
        if a < grass_prob :
            map[n,m,2] = grassTile_id
        else :
            map[n,m,2] = waterTile_id[4]

# 랜덤으로 초기 지형값 세팅

for i in range(0, cellular_iter_num):
    for n in range(0, map_size[0]):
        for m in range(0, map_size[1]):
            try:
                threshold = [map[n-1, m-1, 2], map[n, m-1, 2], map[n+1, m-1, 2],
                             map[n-1, m, 2],map[n, m, 2], map[n+1, m, 2],
                             map[n-1, m+1, 2], map[n, m+1, 2], map[n+1, m+1, 2]]
                if threshold.count(grassTile_id) >= 5:
                    map[n, m, 2] = grassTile_id
            except:
                map[n, m, 2] = grassTile_id



gen = np.random.poisson(lamb, map_size[0]*map_size[1]).reshape(map_size[0], map_size[1])
terrain = np.zeros(map_size)

for n in range(0,map_size[0]):
    for m in range(0,map_size[1]):
        if gen[n, m] > 0:
            for x in range(-5,6):
                for y in range(-5,6):
                    try:
                        prob = random.random()
                        if prob < 0.03 and map[n+x, m+y, 2] == grassTile_id:
                            terrain[n + x, m + y] = woodTerrain_id
                        elif prob < 0.09 and map[n+x, m+y, 2] == grassTile_id:
                            terrain[n + x, m + y] = random.choice([woodTerrain_id, stoneTerrain_id])
                    except:
                        pass
        else:
            pass

        map[n,m,3] = terrain[n,m]







for n in range(0,map_size[0]):
    for m in range(0, map_size[1]):
        map[n, m, 0] = n

        if n%2 == 0: # x축 값이 짝수일 때
            map[n, m, 1] = m*2
        else: # x축 값이 홀수일 때
            map[n, m, 1] = m*2+1

# 맵을 위한 다차원 배열을 만들고, 배열에 좌표계에 따른 좌표값 입력





wb = Workbook()
ws = wb.active
ws.title = "MapData"

for n in range(0, map_size[0]):
    for m in range(0, map_size[1]):
        ws.cell(row=n * map_size[1] + m + 1, column=1, value=map[n, m, 0])
        ws.cell(row=n * map_size[1] + m + 1, column=2, value=map[n, m, 1])
        ws.cell(row=n * map_size[1] + m + 1, column=3, value=map[n, m, 2])
        ws.cell(row=n * map_size[1] + m + 1, column=4, value=map[n, m, 3])

wb.save('Map.xlsx')




heatmap_data = np.zeros(map_size)

for n in range(0,map_size[0]):
    for m in range(0, map_size[1]):
        heatmap_data[n,m] = map[n,m,2]

        if terrain[n,m] != 0:
            heatmap_data[n,m] = terrain[n,m]
        else:
            pass

plt.figure()
heatmap = plt.pcolor(heatmap_data[:, :].T, cmap='tab20', vmax = 20, vmin = 1)
plt.show(heatmap)
#plt.close()



