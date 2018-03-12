import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook

def readTileData(tile_count):
    wb = load_workbook("WorldMap.xlsx")
    tileSheet = wb['WorldMapTile']
    
    tile_id = np.zeros(tile_count)
    
    k=0
    for row in tileSheet.iter_rows(min_row=4, max_row=4-1+tile_count, min_col = 2, max_col=2):
        for cell in row:
            tile_id[k] = cell.value
            k=k+1
    
    
    return tile_id
            

def readTerrainData(terrain_count):
    wb = load_workbook("WorldMap.xlsx")
    terrainSheet = wb['WorldMapTerrain']
    
    terrain_id = np.zeros(terrain_count)

    k=0
    for row in terrainSheet.iter_rows(min_row=4, max_row=4-1+terrain_count, min_col = 2, max_col=2):
        for cell in row:
            terrain_id[k] = cell.value
            k=k+1

    return terrain_id
            

def writeMapData(map_size_x, map_size_y, tile_data_real, terrain_data_real):
    wb = load_workbook("WorldMap.xlsx")
    ws = wb['WorldMapPreset']
    
    k=1
    for n in range(0, map_size_x-1):
        for m in range(0, int(map_size_y/2)-n%2):
            
            ws.cell(row=k+3, column=3, value=n)
            ws.cell(row=k+3, column=4, value=2*m+n%2)
            ws.cell(row=k+3, column=5, value=tile_data_real[n, 2*m+n%2])
            ws.cell(row=k+3, column=6, value=terrain_data_real[n, 2*m+n%2])
            k=k+1
            
    wb.save("WorldMap.xlsx")